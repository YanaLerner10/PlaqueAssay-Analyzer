from pathlib import Path
from openpyxl import Workbook

from reporter_assay_analyzer.stacked_parser import parse_stacked_combined_raw_xlsx


def _write_fake_stacked_combined(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "combined_raw"

    # write ONE timepoint block (0h)
    start = 1
    ws.cell(row=start, column=1, value="0h post transfection")

    # header 1..12 in B..M
    for j in range(1, 13):
        ws.cell(row=start + 1, column=1 + j, value=j)

    # rows A..H with values
    for i, r in enumerate("ABCDEFGH"):
        rr = start + 2 + i
        ws.cell(row=rr, column=1, value=r)
        for c in range(1, 13):
            ws.cell(row=rr, column=1 + c, value=float(c))  # simple numeric pattern

    wb.save(path)


def test_parse_stacked_combined_raw_xlsx(tmp_path: Path):
    p = tmp_path / "combined_raw.xlsx"
    _write_fake_stacked_combined(p)

    tidy = parse_stacked_combined_raw_xlsx(p, sheet_name="combined_raw")
    # 96 wells for one timepoint
    assert len(tidy) == 96
    assert set(tidy.columns) == {"time_h", "well", "value"}
    assert tidy["time_h"].unique().tolist() == [0]
    assert "A1" in tidy["well"].values
    assert "H12" in tidy["well"].values
