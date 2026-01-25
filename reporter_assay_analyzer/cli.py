from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from .mapping import write_mapping_template
from .io import parse_timepoint_hours, read_plate_matrix_xlsx
from .stacked_parser import parse_stacked_combined_raw_xlsx
from .analysis import analyze
from .plots import make_timecourse_plots


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="reporter_assay_analyzer",
        description="Reporter Assay Analyzer - NanoLuc plate time-course processing",
    )
    sub = p.add_subparsers(dest="command", required=True)

    # 1) Generate tidy mapping template CSV
    t = sub.add_parser("make-template", help="Generate a mapping template CSV (A1..H12).")
    t.add_argument("--out", required=True, help="Output path for mapping template CSV.")

    # 2) Combine raw plate exports into one stacked Excel sheet (pretty)
    c = sub.add_parser(
        "combine-raw",
        help="Combine multiple plate files into one Excel as stacked 96-well plates.",
    )
    c.add_argument("--data-dir", required=True, help="Folder with xlsx files (one per timepoint).")
    c.add_argument("--out", required=True, help="Output combined Excel path (e.g., output/combined_raw.xlsx).")

    # 3) Analyze using combined_raw.xlsx + mapping CSV
    a = sub.add_parser(
        "analyze",
        help="Run final analysis using combined_raw.xlsx (stacked plates) + mapping CSV.",
    )
    a.add_argument("--combined", required=True, help="Path to combined_raw.xlsx (stacked format).")
    a.add_argument("--mapping", required=True, help="Mapping CSV file.")
    a.add_argument("--out", required=True, help="Output Excel path (e.g., output/final_analysis.xlsx).")

    # 4) Plot from final_analysis.xlsx
    pp = sub.add_parser("plot", help="Generate time-course plots from final_analysis.xlsx.")
    pp.add_argument("--final", required=True, help="Path to final_analysis.xlsx")
    pp.add_argument("--out-dir", required=True, help="Output directory for plots (e.g., output/plots)")

    return p


def _write_stacked_plates_excel(files: list[Path], out_path: Path) -> None:
    """
    Write a single Excel sheet that contains one 8x12 plate per timepoint,
    stacked vertically, similar to typical plate-reader exports.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "combined_raw"

    start_row = 1

    # sort by timepoint (0h, 1h, 2h...) rather than alphabetical
    files = sorted(files, key=lambda p: parse_timepoint_hours(p.name))

    for f in files:
        t_h = parse_timepoint_hours(f.name)
        title = f"{t_h}h post transfection"

        # 8x12 matrix: index A..H, columns 1..12
        plate = read_plate_matrix_xlsx(f)

        # Title row
        ws.cell(row=start_row, column=1, value=title)

        # Header row: 1..12 in columns B..M
        header_row = start_row + 1
        ws.cell(row=header_row, column=1, value=None)
        for j, col_num in enumerate(range(1, 13), start=2):
            ws.cell(row=header_row, column=j, value=col_num)

        # Data rows: A..H in column A, values in B..M
        for i, row_letter in enumerate("ABCDEFGH", start=0):
            r = start_row + 2 + i
            ws.cell(row=r, column=1, value=row_letter)

            for j, col_num in enumerate(range(1, 13), start=2):
                val = plate.loc[row_letter, col_num]
                ws.cell(row=r, column=j, value=None if pd.isna(val) else float(val))

            # Optional right-side label (matches many exports)
            ws.cell(row=r, column=14, value="Lum")  # column N

        # spacer row
        start_row += 11

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.command == "make-template":
        out_path = Path(args.out)
        write_mapping_template(out_path)
        print(f"✅ Wrote mapping template: {out_path.resolve()}")
        return 0

    if args.command == "combine-raw":
        data_dir = Path(args.data_dir)
        out_path = Path(args.out)

        files = [p for p in data_dir.glob("*.xlsx") if not p.name.startswith("~$")]
        if not files:
            raise SystemExit(f"No .xlsx files found in: {data_dir.resolve()}")

        _write_stacked_plates_excel(files, out_path)
        print(f"✅ Wrote combined raw file (stacked plates): {out_path.resolve()}")
        return 0

    if args.command == "analyze":
        combined_path = Path(args.combined)
        mapping_path = Path(args.mapping)
        out_path = Path(args.out)

        # Parse the stacked combined_raw.xlsx into tidy rows
        tidy = parse_stacked_combined_raw_xlsx(combined_path, sheet_name="combined_raw")

        # Load mapping CSV
        mapping = pd.read_csv(mapping_path)

        # Run analysis
        result = analyze(tidy, mapping)

        # Save
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            result.to_excel(w, index=False, sheet_name="final_analysis")

        print(f"✅ Wrote final analysis: {out_path.resolve()}")
        return 0

    if args.command == "plot":
        final_path = Path(args.final)
        out_dir = Path(args.out_dir)

        final_df = pd.read_excel(final_path, sheet_name="final_analysis")
        make_timecourse_plots(final_df, out_dir)

        print(f"✅ Plots saved to: {out_dir.resolve()}")
        return 0

    parser.error("Unknown command")
    return 2
